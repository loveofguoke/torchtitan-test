# Copyright (c) Meta Platforms, Inc. and affiliates.
# All rights reserved.

"""The only project module that encodes msProbe CLI and output conventions."""

from __future__ import annotations

import csv
import json
from collections import Counter
from pathlib import Path
from typing import Any, Iterable, Sequence

from .artifacts import MindStudioArtifactError, write_json
from .toolchain import resolve_tool_executable


def _msprobe_executable() -> str:
    try:
        return resolve_tool_executable(
            "msprobe",
            environment_variable="TORCHTITAN_MSPROBE",
        )
    except RuntimeError as error:
        raise RuntimeError(
            "msprobe executable was not found; run the MindStudio toolchain "
            "bootstrap or set TORCHTITAN_MSPROBE"
        ) from error


def compare_command(
    *,
    target: Path,
    golden: Path,
    output: Path,
    fuzzy_match: bool = False,
    data_mapping: Path | None = None,
    cell_mapping: Path | None = None,
    diff_analysis: bool = False,
    tensor_log: bool = False,
    xlsx: bool = False,
) -> list[str]:
    executable = _msprobe_executable()
    command = [
        executable,
        "compare",
        "-tp",
        str(target),
        "-gp",
        str(golden),
        "-o",
        str(output),
    ]
    if fuzzy_match:
        command.append("-fm")
    if data_mapping is not None:
        command.extend(("-dm", str(data_mapping)))
    if cell_mapping is not None:
        command.extend(("-cm", str(cell_mapping)))
    if diff_analysis:
        command.append("-da")
    if tensor_log:
        command.append("-tensor_log")
    if xlsx:
        command.append("--xlsx")
    return command


def config_check_compare_command(
    *,
    golden_pack: Path,
    target_pack: Path,
    output: Path,
) -> list[str]:
    executable = _msprobe_executable()
    return [
        executable,
        "config_check",
        "-c",
        str(golden_pack),
        str(target_pack),
        "-o",
        str(output),
    ]


def graph_visualize_command(
    *,
    target: Path,
    golden: Path,
    output: Path,
    fuzzy_match: bool = False,
    overflow_check: bool = False,
    tensor_log: bool = False,
    progress_log: bool = False,
) -> list[str]:
    """Build the public hierarchical graph comparison command."""

    executable = _msprobe_executable()
    command = [
        executable,
        "graph_visualize",
        "-tp",
        str(target),
        "-gp",
        str(golden),
        "-o",
        str(output),
    ]
    if fuzzy_match:
        command.append("-fm")
    if overflow_check:
        command.append("-oc")
    if tensor_log:
        command.append("-tensor_log")
    if progress_log:
        command.append("-progress_log")
    return command


def precheck_command(
    *,
    api_info: Path,
    output: Path,
    device_ids: Sequence[int],
    num_splits: int | None = None,
    jit_compile: bool = False,
    save_error_data: bool = False,
    filter_api: bool = False,
    config_path: Path | None = None,
    resume_csv: Path | None = None,
) -> list[str]:
    """Build the public msProbe acc_check or multi_acc_check command."""

    executable = _msprobe_executable()
    if not device_ids:
        raise ValueError("precheck requires at least one device ID")
    multi = num_splits is not None
    if multi and not 1 <= num_splits <= 64:
        raise ValueError("multi_acc_check num_splits must be in [1, 64]")
    command = [
        executable,
        "multi_acc_check" if multi else "acc_check",
        "-api_info",
        str(api_info),
        "-o",
        str(output),
    ]
    if jit_compile:
        command.append("-j")
    if multi:
        command.extend(("-n", str(num_splits)))
        command.append("-d")
        command.extend(str(device_id) for device_id in device_ids)
    else:
        if len(device_ids) != 1:
            raise ValueError("single acc_check accepts exactly one device ID")
        command.extend(("-d", str(device_ids[0])))
    if save_error_data:
        command.append("-save_error_data")
    if filter_api:
        command.append("-f")
    if config_path is not None:
        command.extend(("-config", str(config_path)))
    if resume_csv is not None:
        command.extend(("-csv_path", str(resume_csv)))
    return command


def precheck_compare_command(
    *,
    npu_details: Path,
    gpu_details: Path,
    output: Path,
) -> list[str]:
    executable = _msprobe_executable()
    return [
        executable,
        "api_precision_compare",
        "-npu",
        str(npu_details),
        "-gpu",
        str(gpu_details),
        "-o",
        str(output),
    ]


def find_precheck_details(directory: Path) -> Path:
    matches = sorted(directory.rglob("accuracy_checking_details_*.csv"))
    if len(matches) != 1:
        raise MindStudioArtifactError(
            f"expected one accuracy_checking_details CSV under {directory}, "
            f"found {len(matches)}"
        )
    return matches[0]


def find_dump_compare_input(artifact_directory: Path, step: int) -> Path:
    step_directory = artifact_directory / "official" / f"step{step}"
    if not step_directory.is_dir():
        raise MindStudioArtifactError(
            f"msProbe step directory is missing: {step_directory}"
        )
    rank_directories = sorted(
        path for path in step_directory.iterdir() if path.is_dir()
    )
    if not rank_directories:
        raise MindStudioArtifactError(f"no msProbe ranks found: {step_directory}")
    if len(rank_directories) == 1:
        dump_path = rank_directories[0] / "dump.json"
        if not dump_path.is_file():
            raise MindStudioArtifactError(f"msProbe dump is missing: {dump_path}")
        return dump_path
    for rank_directory in rank_directories:
        if not (rank_directory / "dump.json").is_file():
            raise MindStudioArtifactError(
                f"msProbe dump is missing: {rank_directory / 'dump.json'}"
            )
    return step_directory


def _normalized_status(value: str) -> str | None:
    normalized = value.strip().lower().replace("_", " ")
    if normalized in {"pass", "passed", "true", "yes", "通过", "一致"}:
        return "pass"
    if normalized in {"warning", "warn", "警告"}:
        return "warning"
    if normalized in {
        "error",
        "fail",
        "failed",
        "false",
        "no",
        "不通过",
        "不一致",
    }:
        return "error"
    if normalized in {"skip", "skipped"}:
        return "skip"
    return None


_STATUS_FIELDS = {
    "result",
    "status",
    "compare result",
    "forward test success",
    "backward test success",
    "是否一致",
    "比对结果",
}


def _status_values(values: Iterable[Any]) -> tuple[Counter[str], bool]:
    counts: Counter[str] = Counter()
    has_unknown_value = False
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        status = _normalized_status(text)
        if status is None:
            has_unknown_value = True
        else:
            counts[status] += 1
    return counts, bool(counts) and not has_unknown_value


def _csv_statuses(path: Path) -> tuple[Counter[str], bool]:
    result: Counter[str] = Counter()
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as stream:
        reader = csv.DictReader(stream)
        if not reader.fieldnames:
            return result, False
        status_fields = [
            field
            for field in reader.fieldnames
            if field.strip().lower() in _STATUS_FIELDS
        ]
        if not status_fields:
            return result, False
        values: list[Any] = []
        for row in reader:
            for field in status_fields:
                values.append(row.get(field, ""))
    return _status_values(values)


def _xlsx_statuses(path: Path) -> tuple[Counter[str], bool]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return Counter(), False
    values: list[Any] = []
    found_status_field = False
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                continue
            status_indices = [
                index
                for index, value in enumerate(header)
                if str(value or "").strip().lower() in _STATUS_FIELDS
            ]
            found_status_field = found_status_field or bool(status_indices)
            for row in rows:
                for index in status_indices:
                    values.append(row[index] if index < len(row) else "")
    finally:
        workbook.close()
    if not found_status_field:
        return Counter(), False
    return _status_values(values)


def summarize_official_results(output: Path) -> dict[str, Any]:
    """Summarize official verdict fields without recomputing any metric."""

    files = sorted(path for path in output.rglob("*") if path.is_file())
    status_counts: Counter[str] = Counter()
    decisive_files: list[str] = []
    parsed_files: list[str] = []
    unparsed_files: list[str] = []
    for path in files:
        suffix = path.suffix.lower()
        if suffix not in {".csv", ".xlsx"}:
            continue
        relative = path.relative_to(output).as_posix()
        decisive_files.append(relative)
        counts, fully_parsed = (
            _csv_statuses(path) if suffix == ".csv" else _xlsx_statuses(path)
        )
        status_counts.update(counts)
        if fully_parsed:
            parsed_files.append(relative)
        else:
            unparsed_files.append(relative)
    verdict = "unparsed"
    if status_counts and not unparsed_files:
        verdict = (
            "error"
            if status_counts["error"]
            else "warning"
            if status_counts["warning"]
            else "warning"
            if status_counts["skip"]
            else "pass"
        )
    summary = {
        "schema": "torchtitan.glm5_2.msprobe_official_summary",
        "schema_version": 2,
        "verdict": verdict,
        "status_counts": dict(sorted(status_counts.items())),
        "decisive_result_files": decisive_files,
        "parsed_decisive_result_files": parsed_files,
        "unparsed_decisive_result_files": unparsed_files,
        "parsed_result_files": parsed_files,
        # Retain the original field for report readers created before XLSX
        # config-check output was integrated.
        "parsed_csv": parsed_files,
        "official_files": [path.relative_to(output).as_posix() for path in files],
        "note": (
            "The project only aggregates verdict fields emitted by msProbe; "
            "it does not recompute or replace official accuracy metrics."
        ),
    }
    write_json(output / "official_summary.json", summary)
    return summary


def validate_compile_report(path: Path) -> dict[str, Any]:
    """Validate that a PrecisionChecker CSV contains decisive module rows.

    A single-pass report always contains a synthetic ``LOSS`` row whose status
    can be PASS even though no eager whole-model loss exists.  That row, and a
    CSV header on its own, are not evidence that any wrapped module ran.
    """

    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as stream:
        reader = csv.DictReader(stream)
        field_names = {
            str(field).strip().lower(): field for field in (reader.fieldnames or ())
        }
        required = {"module_name", "check_type", "status"}
        missing = sorted(required - set(field_names))
        if missing:
            raise MindStudioArtifactError(
                f"PrecisionChecker CSV is missing required columns {missing}: {path}"
            )

        forward_rows = 0
        backward_rows = 0
        module_names: set[str] = set()
        status_counts: Counter[str] = Counter()
        for row in reader:
            module_name = str(row.get(field_names["module_name"], "") or "").strip()
            check_type = str(
                row.get(field_names["check_type"], "") or ""
            ).strip().lower()
            status = str(row.get(field_names["status"], "") or "").strip().upper()
            if not module_name or module_name.upper() == "LOSS":
                continue
            if status in {"SKIP", "IGNORED", ""}:
                continue
            if status not in {"PASS", "FAIL", "WARN", "WARNING"}:
                continue
            if check_type == "fwd_output":
                forward_rows += 1
            elif check_type in {"grad_input", "grad_output"}:
                backward_rows += 1
            else:
                continue
            module_names.add(module_name)
            status_counts[status.lower()] += 1

    if not forward_rows or not backward_rows:
        raise MindStudioArtifactError(
            "PrecisionChecker report must contain non-skipped module forward "
            f"and backward rows; found forward={forward_rows}, "
            f"backward={backward_rows}: {path}"
        )
    return {
        "forward_rows": forward_rows,
        "backward_rows": backward_rows,
        "module_names": sorted(module_names),
        "status_counts": dict(sorted(status_counts.items())),
    }


def aggregate_compile_csv(
    rank_csv_files: Sequence[Path],
    *,
    output: Path,
    scenario: str,
) -> Path:
    if not rank_csv_files:
        raise MindStudioArtifactError("no per-rank PrecisionChecker CSV files found")
    output.parent.mkdir(parents=True, exist_ok=True)
    header: list[str] | None = None
    with output.open("w", encoding="utf-8", newline="") as destination:
        writer = csv.writer(destination)
        for rank_csv in sorted(rank_csv_files):
            rank_text = rank_csv.stem.removeprefix("precision_rank")
            with rank_csv.open(
                "r", encoding="utf-8-sig", errors="replace", newline=""
            ) as source:
                reader = csv.reader(source)
                current_header = next(reader, None)
                if current_header is None:
                    continue
                if header is None:
                    header = current_header
                    writer.writerow(("rank", "scenario", *header))
                elif current_header != header:
                    raise MindStudioArtifactError(
                        f"PrecisionChecker CSV header mismatch: {rank_csv}"
                    )
                for row in reader:
                    writer.writerow((rank_text, scenario, *row))
    return output


def write_compare_invocation(path: Path, command: Iterable[str]) -> None:
    write_json(path, {"command": list(command)})
